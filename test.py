import unittest
from datetime import datetime

from convert import convert_csv_to_excel
import openpyxl as xl
import csv
from itertools import zip_longest
from compare import sort_values, ValueNode, make_sorted_sheet, compare_files, is_number, is_date
import logging
from dateutil.parser import parse


class TestConvert(unittest.TestCase):
    def test_convert_csv(self):
        csv_path = r"tests\left.csv"
        excel_file = convert_csv_to_excel(csv_path)
        wb = xl.load_workbook(excel_file)
        ws = wb.active
        with open(csv_path, newline='') as csv_file:
            reader = csv.reader(csv_file, delimiter=',', quotechar='"')
            cells_checked = 0
            for (excel_row, csv_row) in zip_longest(ws.rows, reader):
                for i in range(0, len(excel_row)):
                    self.assertEqual(excel_row[i].value, csv_row[i], "excel value differs from csv")
                    cells_checked += 1
            self.assertGreater(cells_checked, 0)  # greater than zero


class TestExcel(unittest.TestCase):
    def setUp(self):

        self.left_xlsx = r"tests\left.xlsx"
        self.right_xlsx = r"tests\right.xlsx"
        self.left_wb = xl.load_workbook(self.left_xlsx)
        self.right_wb = xl.load_workbook(self.right_xlsx)

        self.left_sheet = self.left_wb.worksheets[0]
        self.right_sheet = self.right_wb.worksheets[0]
        self.val = sort_values(self.left_sheet, self.right_sheet, 1, True)

    def test_sort(self):
        expected_values = [(2, 2, "Row 1"), (3, 3, "Row 2"), (5, None, "Row 3"), (4, 4, "Row 4")]
        test_list = [ValueNode(f, r, v) for (f, r, v) in expected_values]
        check_count = 0
        for (i, x) in zip_longest(self.val, test_list):
            self.assertEqual(i, x, "test value does not match expected value")
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero

    def test_make_sorted_sheet(self):
        wb = xl.Workbook()
        sheet_name = self.left_wb.sheetnames[0]
        left_sheet_name = 'left_' + sheet_name
        left_sheet = make_sorted_sheet(wb, self.left_sheet, self.val, left_sheet_name, 'left', True)
        expected_values = [(1, 1, "Header"), (2, 2, "Row 1"), (3, 3, "Row 2"), (5, None, "Row 3"), (4, 4, "Row 4")]
        check_count = 0
        for row in range(1, left_sheet.max_row + 1):
            self.assertEqual(left_sheet.cell(row=row, column=1).value, expected_values[row - 1][2])
            check_count += 1
        self.assertGreater(check_count, 0)  # greater than zero

    def test_compare_files_xlsx(self):
        compare_files(self.left_xlsx, self.right_xlsx, r"tests\output.xlsx", open_on_finish=False,
                      sort_column=1, compare_type="sorted", sheet_matching="order")
        expected_values = [
            ["Header", "Header", "Same", "Col A", "Col A", "Same", "Col B", "Col B", "Same"],
            ["Row 1", "Row 1", "Same", "1", 1, 0, "2", 3, 1],
            ["Row 2", "Row 2", "Same", "z", "z", "Same", "Q", "W", "Different"],
            ["Row 3", None, "Different", "extra", None, "Different", "row", None, "Different"],
            ["Row 4", "Row 4", "Same", "1/1/2019", datetime(2019, 1, 1), "Different", "2/2/2012",
             datetime(2012, 2, 1), "Different"]
        ]

        wb = xl.load_workbook(r"tests\output.xlsx")
        ws = wb.worksheets[2]  # third worksheet is diff
        for col in range(1, ws.max_column + 1):
            for row in range(1, ws.max_row + 1):

                value = ws.cell(row=row, column=col).value
                expected_value = expected_values[row - 1][col - 1]
                if is_number(value) and is_number(expected_value):
                    self.assertEqual(float(value), float(expected_value))
                else:
                    self.assertEqual(str(value), str(expected_value))

    def test_compare_files_csv(self):
        compare_files(r"tests\left.csv", r"tests\right2.csv", r"tests\output2.xlsx", open_on_finish=False,
                      sort_column=1, compare_type="sorted", sheet_matching="order")
        expected_values = [
            ["Header", "Header", "Same", "Col A", "Col A", "Same", "Col B", "Col B", "Same"],
            ["Row 1", "Row 1", "Same", "1", 1, 0, "2", 3, 1],
            ["Row 2", "Row 2", "Same", "z", "z", "Same", "Q", "W", "Different"],
            ["Row 3", None, "Different", "extra", None, "Different", "row", None, "Different"],
            ["Row 4", "Row 4", "Same", "1/1/2019", "1/1/2019", "Same", "2/2/2012",
             "2/1/2012", "Different"]
        ]
        wb = xl.load_workbook(r"tests\output2.xlsx")
        ws = wb.worksheets[2]  # third worksheet is diff
        for col in range(1, ws.max_column + 1):
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                expected_value = expected_values[row - 1][col - 1]
                #logging.debug("row {}, col {}: {} vs {}".format(row, col, value, expected_value))
                if is_number(value) and is_number(expected_value):
                    self.assertEqual(float(value), float(expected_value))
                elif is_date(value) and is_date(expected_value):
                    self.assertEqual(parse(value), parse(expected_value))
                else:
                    self.assertEqual(str(value), str(expected_value))
