import argparse
import logging
import openpyxl as xl
from openpyxl.styles import PatternFill
import os
from collections import namedtuple
from dateutil.parser import parse
from convert import convert_csv_to_excel

logging.basicConfig(level=logging.DEBUG, format='%(asctime)-15s %(message)s')
ValueNode = namedtuple('ValueNode', ['left_row', 'right_row', 'value'])  # object to store left row #, right #, value


def make_sorted_sheet(workbook, sheet, sorted_values, new_sheet_name, left_or_right, has_header=True):
    new_sheet = workbook.create_sheet(title=new_sheet_name)
    max_col = sheet.max_column

    logging.info("make sorted sheet: '{}', '{}'".format(new_sheet_name, left_or_right))
    if has_header:
        header_values = []
        for c in range(1, max_col + 1):
            header_values.append(sheet.cell(row=1, column=c).value)
        new_sheet.append(header_values)

    for v in sorted_values:
        if left_or_right == "left":
            row_id = v.left_row
        else:
            row_id = v.right_row

        # if a row id exists for this value & sheet, copy row data to new row in new sheet
        row_values = []
        if row_id is not None:  # if no row id, row_values will be empty
            for c in range(1, max_col + 1):
                row_values.append(sheet.cell(row=row_id, column=c).value)
        else:
            row_values = [None]
        new_sheet.append(row_values)  # append to new sheet
    return new_sheet


def compare_files(left_path, right_path, output_path, threshold=0.001, open_on_finish=False, sort_column=None,
                  compare_type="default", has_header=True, sheet_matching="name"):
    logging.info(
        "Comparing '{}' vs '{}' with threshold = '{}', sort column = '{}', compare type='{}'".format(
            left_path, right_path, threshold, sort_column, compare_type))

    # check file extension if valid.  If not, convert CSV to
    logging.info("validating file types: '{}', '{}'".format(left_path, right_path))
    left_path = is_file_extension_valid(left_path)
    right_path = is_file_extension_valid(right_path)

    # load workbook into excel library
    logging.info("loading excel files: '{}', '{}'".format(left_path, right_path))
    left_wb = xl.load_workbook(filename=left_path)
    right_wb = xl.load_workbook(filename=right_path)
    output_wb = xl.Workbook()

    # get sheet names
    logging.info("get sheet names")
    left_sheets = left_wb.sheetnames
    right_sheets = right_wb.sheetnames
    output_sheets = output_wb.sheetnames

    if sheet_matching == "name":
        sheets_to_process = [(sheet, sheet) for sheet in left_sheets if sheet in right_sheets]
    else:
        sheets_to_process = list(map(lambda i, j: (i, j), left_sheets, right_sheets))

    logging.info("sheet match style: '{}', sheets to process: {}".format(sheet_matching, sheets_to_process))

    if len(sheets_to_process) > 0:  # remove default sheet
        for sheet in output_sheets:
            output_wb.remove(output_wb[sheet])
    else:
        raise ValueError("No sheets were found for processing.  Check sheet_matching parameter is set correctly " +
                         "(e.g. name or order)")

    for (i, j) in sheets_to_process:
        left_sheet = left_wb[i]
        right_sheet = right_wb[j]

        if compare_type == "sorted":
            logging.info("sorting sheets prior to comparison: ({},{})".format(i, j))

            sorted_values = sort_values(left_sheet, right_sheet, sort_column, has_header)
            logging.info("values sorted.  Sorting left sheet")
            left_sheet = make_sorted_sheet(output_wb, left_sheet, sorted_values, 'left_' + i, 'left', has_header)
            logging.info("left sorted.  Sorting right sheet")
            right_sheet = make_sorted_sheet(output_wb, right_sheet, sorted_values, 'right_' + j, 'right',
                                            has_header)
        output_sheet_name = i if sheet_matching == "name" or i == j else "{} v {}".format(i, j)
        output_sheet = output_wb.create_sheet(output_sheet_name)

        logging.info("comparing sheets: ({},{})".format(i, j))
        compare_sheet(left_sheet, right_sheet, output_sheet, threshold)

    logging.info("saving to file: '{}'".format(output_path))
    output_wb.save(output_path)

    logging.info("save complete")
    if open_on_finish:
        path_to_open = '"' + output_path + '"'
        logging.info("opening file".format(path_to_open))
        os.system(path_to_open)  # use OS command line to open file.  This works on Windows


def is_file_extension_valid(file_path):
    if is_extension(file_path, '.csv'):
        file_path = convert_csv_to_excel(file_path)
    elif not is_extension(file_path, ".xlsx"):
        raise ValueError("file extension for {} is not xlsx or csv.  file cannot be processed.".format(file_path))
    return file_path


def sort_values(left, right, sort_column, has_header=False):
    if sort_column is None:  # e.g., if not none
        return

    starting_row = 1 if has_header is False else 2

    # get list of named tuples. left side populates x.  right populates y.  merge later.
    x = [ValueNode(n, None, left.cell(row=n, column=sort_column).value) for n in
         range(starting_row, left.max_row + 1)]  # x = left
    y = [ValueNode(None, n, right.cell(row=n, column=sort_column).value) for n in
         range(starting_row, right.max_row + 1)]  # y = right

    logging.debug("starting_row for sort: {}".format(starting_row))
    # if there are any values that are not numbers, convert all to string
    if (
            len([i for i in x if not is_number(i.value)]) > 0 or
            len([i for i in y if not is_number(i.value)]) > 0):
        x = sorted([ValueNode(i.left_row, i.right_row, str(i.value)) for i in x], key=lambda tup: tup[2])
        y = sorted([ValueNode(i.left_row, i.right_row, str(i.value)) for i in y], key=lambda tup: tup[2])
    else:  # otherwise, convert all to number
        x = sorted([ValueNode(i.left_row, i.right_row, float(i.value)) for i in x], key=lambda tup: tup[2])
        y = sorted([ValueNode(i.left_row, i.right_row, float(i.value)) for i in y], key=lambda tup: tup[2])

    i = j = 0

    z = []  # z is combined list
    while i < len(x) and j < len(y):

        f = x[i].value
        r = y[j].value

        # compare left and right values.  if both values match, combine into a single tuple.  otherwise, take only one
        if f == r:
            temp = ValueNode(x[i].left_row, y[j].right_row, f)
            z.append(temp)
            i += 1
            j += 1
        elif f < r:  # left side has number lower than right side
            z.append(x[i])  # add node from left side since right side is None
            i += 1
        elif r < f:  # right side has number lower than left side
            z.append(y[j])  # add node from right side since left is None
            j += 1
        else:
            logging.warning("encountered a case where no comparison can be made {} vs {}".format(f, r))

    while i < len(x):
        z.append(x[i])
        i += 1
    while j < len(y):
        z.append(y[j])
        j += 1

    return z


def compare_sheet(left_sheet, right_sheet, output_sheet, threshold):
    max_col = max(left_sheet.max_column, right_sheet.max_column)
    max_row = max(left_sheet.max_row, right_sheet.max_row)

    columns_per_value = 3  # each comparison takes up 3 rows
    left_offset = 0
    right_offset = 1
    diff_offset = 2

    starting_row = 1

    for col in range(1, max_col + 1):
        output_column = (col - 1) * columns_per_value + 1  # 1-based column count, offset by columns per value
        for row in range(starting_row, max_row + 1):
            left_cell = left_sheet.cell(row=row, column=col)
            right_cell = right_sheet.cell(row=row, column=col)

            diff_value = cell_difference(left_cell, right_cell)

            output_sheet.cell(row=row, column=output_column + left_offset).value = left_cell.value  # output left
            output_sheet.cell(row=row, column=output_column + right_offset).value = right_cell.value  # output right
            output_sheet.cell(row=row, column=output_column + diff_offset).value = diff_value  # output diff
            apply_style(output_sheet.cell(row=row, column=output_column + diff_offset), threshold)

    return output_sheet


def cell_difference(left_cell, right_cell):
    return value_difference(left_cell.value, right_cell.value)


def value_difference(left, right):
    if is_number(left) and is_number(right):
        diff_value = float(right) - float(left)  # numbers are subtracted
    elif is_date(left) and is_date(right):
        diff_value = "Same" if parse(right) == parse(left) else "Different"  # date comparison
    else:
        diff_value = "Same" if right == left else "Different"  # non-number comparison
    return diff_value


def is_number(my_value):
    try:
        float(my_value)
        return True
    except Exception:
        return False


def is_date(my_value):
    try:
        parse(my_value)
        return True
    except Exception:
        return False


def apply_style(cell, threshold):
    same_color = "93f277"
    different_color = "edb26f"
    pattern_same = PatternFill(start_color=same_color, fill_type="solid")
    pattern_diferent = PatternFill(start_color=different_color, fill_type="solid")
    if is_number(cell.value) and abs(cell.value) <= threshold:
        cell.fill = pattern_same  # under threshold
    elif cell.value == "Same":
        cell.fill = pattern_same  # match
    else:
        cell.fill = pattern_diferent  # under threshold or different


def is_extension(file_path, extension):
    (file_name, file_extension) = os.path.splitext(file_path)
    return file_extension.lower() == extension


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Compare two Excel (XLSX) files sheet by sheet.  " +
                                                 "Create a new excel file with values side by side along with " +
                                                 "differences.")
    parser.add_argument("left", help="Path to first file for comparison.  Can be CSV or XLSX.  In output file, these " +
                                     "values will be on left")

    parser.add_argument("right", help="Path to second file for comparison.  Can be CSV or XLSX.  In output file, " +
                                      " these values will be on right")
    parser.add_argument("output", help="Path to output file.  If file exists it will be overwritten.  If " +
                                       "compare_type is 'sorted' then it will contain copies of data from original " +
                                       "files as well as the values side by side in a combined sheet.")
    parser.add_argument("--threshold", '-t', type=float, default=0.001,
                        help="threshold for numeric values to be considered different.  e.g. when threshold = 0.01 " +
                             "if left and right values are closer than 0,01 then consider the same.  Mainly affects " +
                             "coloring of difference column for numeric values")
    parser.add_argument("--open", '-p', type=bool, default=True, help="if true, open output file on completion " +
                                                                      "using os.system.  Output file path must " +
                                                                      "resolve to a file.  Adds quotes around file " +
                                                                      "name so that paths with spaces can resolve" +
                                                                      "on windows machines.")
    parser.add_argument("--compare_type", '-c', default="default",
                        help="if set to 'sorted', the comparison tool will attempt to line up each side based on " +
                             "the values of sort_column specified.  'default' is a cell-by-cell comparison.")
    parser.add_argument("--sort_column", "-s", type=int, default=None,
                        help="numeric offset (1-based) of column to use for sorting.  E.g. a primary key. " +
                             "if compare type is 'sorted', this column will be used to sort and line up each side")
    parser.add_argument("--has_header", "-d", default=True,
                        help="if sheets have headers, set to True so the headers can be excluded from comparison")
    parser.add_argument("--sheet_matching", "-m", default="order", help="can be either 'name' or 'order'.  If name, " +
                                                                        "only sheets with the same name are " +
                                                                        "compared. " +
                                                                        "if order, sheets are compared in order. " +
                                                                        "E.g. 1st sheet vs 1st sheet.")
    parser.add_argument("--convert_csv", "-v", default=True, help="if True, convert csv files to xlsx")
    args = parser.parse_args()
    if args.compare_type == "sorted":
        if not is_number(args.sort_column):
            parser.error("sort column must be a number if compare type is sorted")

    compare_files(args.left, args.right, args.output, args.threshold, args.open, args.sort_column, args.compare_type,
                  args.has_header, args.sheet_matching)
